# scrape_basic_parts_playwright.py
import os
import re
import time
from urllib.parse import urljoin
from typing import Optional, Dict, List, Tuple, Union

from playwright.sync_api import sync_playwright, expect, TimeoutError as PWTimeout
from openpyxl import Workbook

URL_LIST = "https://jlcpcb.com/parts/basic_parts"
OUT_XLSX_BASE = "jlcpcb_basic_parts"

RESULTS_RE = re.compile(r"Results remaining:\s*([\d,]+)", re.I)


def ts() -> str:
    return time.strftime("%H:%M:%S")


def log(msg: str) -> None:
    print(f"[{ts()}] {msg}")


def norm(s: str) -> str:
    return " ".join((s or "").split())


def fmt_dur(seconds: float) -> str:
    if seconds < 0:
        seconds = 0
    s = int(seconds)
    h = s // 3600
    m = (s % 3600) // 60
    sec = s % 60
    if h > 0:
        return f"{h:d}:{m:02d}:{sec:02d}"
    return f"{m:d}:{sec:02d}"


def parse_int_from_digits(s: str) -> Union[int, str]:
    m = re.search(r"[\d,]+", s or "")
    if not m:
        return norm(s)
    try:
        return int(m.group(0).replace(",", ""))
    except Exception:
        return norm(s)


def get_results_remaining(page) -> Optional[int]:
    try:
        loc = page.locator("text=/Results remaining:/").first
        txt = norm(loc.inner_text(timeout=3000))
        m = RESULTS_RE.search(txt)
        return int(m.group(1).replace(",", "")) if m else None
    except Exception:
        return None


def wait_results_change(page, old: Optional[int], timeout_s: float = 15.0) -> Optional[int]:
    t0 = time.time()
    while time.time() - t0 < timeout_s:
        cur = get_results_remaining(page)
        if cur is not None and old is not None and cur != old:
            return cur
        time.sleep(0.25)
    return get_results_remaining(page)


def wait_table_ready(page, timeout_ms: int = 60000) -> None:
    page.wait_for_function(
        "() => document.querySelectorAll('.el-table__body-wrapper tbody tr').length > 0",
        timeout=timeout_ms,
    )
    mask = page.locator(".el-loading-mask").first
    try:
        if mask.count():
            mask.wait_for(state="hidden", timeout=timeout_ms)
    except PWTimeout:
        pass


def set_parts_type_basic(page) -> None:
    block = page.locator("xpath=//*[contains(normalize-space(.),'Parts Type')]/ancestor::div[1]").first
    label = block.locator("label.el-checkbox", has_text=re.compile(r"\bBasic\b", re.I)).first
    cb = label.locator("input[type='checkbox']").first

    label.wait_for(state="visible", timeout=15000)
    if not cb.is_checked():
        label.click(timeout=5000)
    expect(cb).to_be_checked()


def click_apply(page) -> None:
    btn = page.get_by_role("button", name="Apply").first
    btn.wait_for(state="visible", timeout=15000)
    expect(btn).to_be_enabled(timeout=30000)
    btn.click(timeout=10000)


def goto_next_page_list(page) -> bool:
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(0.2)

    next_btn = page.locator(".el-pagination button.btn-next").first
    if next_btn.count() == 0:
        return False
    try:
        if next_btn.is_disabled():
            return False
    except Exception:
        if next_btn.get_attribute("disabled") is not None:
            return False

    active = page.locator(".el-pagination .el-pager li.number.active").first
    cur = norm(active.inner_text())
    cur_i = int(cur) if cur.isdigit() else None

    next_btn.click(timeout=5000)
    if cur_i is not None:
        expect(active).to_have_text(str(cur_i + 1), timeout=30000)

    wait_table_ready(page, timeout_ms=60000)
    return True


def scrape_list_basic_parts(page) -> Tuple[List[Dict], Dict]:
    wait_table_ready(page)

    payload = page.evaluate(
        r"""
() => {
  const rows = Array.from(document.querySelectorAll(".el-table__body-wrapper tbody tr"));
  let rows_total = rows.length;
  let rows_with_link = 0;
  let kept = [];

  for (const row of rows) {
    const rowText = (row.innerText || "").replace(/\s+/g, " ").trim();
    const codeMatch = rowText.match(/\bC\d{3,}\b/);
    const code = codeMatch ? codeMatch[0] : "";

    const a = row.querySelector("a[href*='/partdetail/']");
    const href = a ? (a.getAttribute("href") || "") : "";
    if (href) rows_with_link++;

    if (code && href) kept.push({ code, href });
  }

  return {
    stats: { rows_total, rows_with_link, rows_kept: kept.length },
    items: kept
  };
}
"""
    )

    stats = payload.get("stats", {})
    items_raw = payload.get("items", [])

    items: List[Dict] = []
    for x in items_raw:
        link = urljoin(page.url, x["href"])
        items.append({"Code": x["code"], "LinkURL": link})

    return items, stats


def extract_detail_data(page_detail) -> Dict[str, str]:
    return page_detail.evaluate(
        r"""
() => {
  function xFirst(xpath) {
    try { return document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue; }
    catch (e) { return null; }
  }

  function labelValueText(label) {
    const node = xFirst(`//*[normalize-space()='${label}']`);
    if (!node) return '';
    let v = node.nextElementSibling;
    if (!v && node.parentElement && node.parentElement.children && node.parentElement.children.length >= 2) {
      v = node.parentElement.children[1];
    }
    const t = v ? (v.innerText || '') : '';
    return t.replace(/\s+/g,' ').trim();
  }

  function labelValueHref(label) {
    const node = xFirst(`//*[normalize-space()='${label}']`);
    if (!node) return '';
    // чаще всего ссылка в следующем sibling или в контейнере строки
    const cand1 = node.nextElementSibling ? node.nextElementSibling.querySelector("a[href]") : null;
    const cand2 = node.parentElement ? node.parentElement.querySelector("a[href]") : null;
    const a = cand1 || cand2;
    if (!a) return '';
    let href = a.href || a.getAttribute('href') || '';
    href = (href || '').trim();
    if (!href) return '';
    try { href = new URL(href, location.origin).href; } catch(e) {}
    return href;
  }

  // Package/Description
  const pkg  = labelValueText('Package');
  const desc = labelValueText('Description');

  // In Stock
  const body = document.body ? (document.body.innerText || '') : '';
  const m = /In\s*Stock:\s*([\d,]+)/i.exec(body);
  const instock = m ? m[1] : '';

  // Breadcrumb: берём список <a> рядом с "All Components"
  let crumbs = [];
  const allLinks = Array.from(document.querySelectorAll("a"));
  const ac = allLinks.find(a => (a.textContent || '').replace(/\s+/g,' ').trim().startsWith('All Components'));
  if (ac) {
    let node = ac.parentElement;
    for (let up=0; up<6 && node; up++) {
      const c = node.querySelectorAll("a");
      if (c && c.length >= 2) break;
      node = node.parentElement;
    }
    if (node) {
      const arr = Array.from(node.querySelectorAll("a"))
        .map(a => (a.textContent || '').replace(/\s+/g,' ').trim())
        .map(t => t.replace(/\s*\/\s*$/,'').trim())
        .filter(Boolean);
      // drop leading "All Components"
      if (arr.length && arr[0].toLowerCase().startsWith('all components')) arr.shift();
      crumbs = arr;
    }
  }
  const categoryPath = crumbs.join(' / ');
  const type = crumbs.length >= 2 ? crumbs.slice(-2).join(' / ') : (crumbs[0] || '');

  // Datasheet: сначала через label, потом fallback regex по HTML
  let ds = labelValueHref('Datasheet');

  if (!ds) {
    const html = document.documentElement ? (document.documentElement.innerHTML || '') : '';
    const mm = /\/api\/file\/downloadByFileSystemAccessId\/\d+/i.exec(html);
    if (mm) {
      try { ds = new URL(mm[0], location.origin).href; } catch(e) { ds = location.origin + mm[0]; }
    }
  }

  return { pkg, desc, instock, ds, categoryPath, type };
}
"""
    )


def try_datasheet_via_download_event(page_detail) -> str:
    """
    Если DOM/regex не дали ссылку, пробуем кликнуть 'Datasheet -> Download' и поймать download.url,
    после чего сразу cancel(), чтобы ничего не сохранять.
    """
    try:
        ds_row = page_detail.locator("xpath=//*[normalize-space()='Datasheet']/following-sibling::*[1]").first
        btn = ds_row.get_by_text(re.compile(r"Download", re.I)).first
        if btn.count() == 0:
            btn = page_detail.get_by_text(re.compile(r"^Download$", re.I)).first

        with page_detail.expect_download(timeout=4000) as di:
            btn.click(timeout=3000)
        dl = di.value
        url = getattr(dl, "url", "") or ""
        try:
            dl.cancel()  # supported in Playwright
        except Exception:
            pass
        try:
            dl.delete()
        except Exception:
            pass
        return url.strip()
    except Exception:
        return ""


def enrich_one_detail(
    page_detail,
    item: Dict,
    min_interval_s: float,
    retries: int,
    breadcrumb_wait_s: float,
    datasheet_click_fallback: bool,
) -> Dict:
    url = item["LinkURL"]
    last_err = None

    for attempt in range(1, retries + 1):
        t_iter = time.perf_counter()
        try:
            page_detail.goto(url, wait_until="commit", timeout=120000)

            # Stage A: ждём базовые поля (Package/Description/In Stock) до ~8s
            data = None
            for _ in range(60):  # 60 * 150ms = 9s
                d = extract_detail_data(page_detail)
                if d.get("pkg") and d.get("desc") and d.get("instock"):
                    data = d
                    break
                page_detail.wait_for_timeout(150)
            if data is None:
                data = extract_detail_data(page_detail)

            # Stage B: коротко подождать breadcrumb/datasheet, если пусто
            t_b = time.perf_counter()
            while time.perf_counter() - t_b < breadcrumb_wait_s:
                need_cat = not (data.get("categoryPath") or data.get("type"))
                need_ds = not data.get("ds")
                if not need_cat and not need_ds:
                    break
                page_detail.wait_for_timeout(150)
                d2 = extract_detail_data(page_detail)
                # мерджим только если появилось
                if (d2.get("categoryPath") or d2.get("type")) and need_cat:
                    data["categoryPath"] = d2.get("categoryPath", "")
                    data["type"] = d2.get("type", "")
                if d2.get("ds") and need_ds:
                    data["ds"] = d2.get("ds", "")

            # Заполняем
            item["Package"] = data.get("pkg", "")
            item["Description"] = data.get("desc", "")
            item["CategoryPath"] = data.get("categoryPath", "")
            item["Type"] = data.get("type", "")

            inst = data.get("instock", "")
            item["InStock"] = parse_int_from_digits(inst) if inst else ""

            ds_url = norm(data.get("ds", ""))
            if not ds_url and datasheet_click_fallback:
                ds_url = try_datasheet_via_download_event(page_detail)
            item["DatasheetURL"] = ds_url

            item.pop("_detail_error", None)

            elapsed = time.perf_counter() - t_iter
            remain = min_interval_s - elapsed
            if remain > 0:
                page_detail.wait_for_timeout(int(remain * 1000))

            return item

        except Exception as e:
            last_err = e
            page_detail.wait_for_timeout(700 + 400 * attempt)

    item["_detail_error"] = str(last_err)[:180] if last_err else "unknown"
    return item


def save_excel(rows: List[Dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BasicParts"

    headers = ["Code", "Type", "Package", "In Stock", "Description", "Link", "Datasheet", "_detail_error", "CategoryPath"]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append([
            r.get("Code", ""),
            r.get("Type", ""),
            r.get("Package", ""),
            r.get("InStock", ""),
            r.get("Description", ""),
            "",
            "",
            r.get("_detail_error", ""),
            r.get("CategoryPath", ""),
        ])
        row_idx = ws.max_row

        link_url = r.get("LinkURL", "")
        c_link = ws.cell(row=row_idx, column=6)
        if link_url:
            c_link.value = "Link"
            c_link.hyperlink = link_url
            c_link.style = "Hyperlink"

        ds_url = r.get("DatasheetURL", "")
        c_ds = ws.cell(row=row_idx, column=7)
        if ds_url:
            c_ds.value = "Link"
            c_ds.hyperlink = ds_url
            c_ds.style = "Hyperlink"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 90
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 60

    wb.save(out_path)


def main():
    headless = os.environ.get("HEADLESS", "1") != "0"
    min_interval = float(os.environ.get("DETAIL_DELAY", "2"))  # минимальная длительность итерации на детальной странице
    retries = int(os.environ.get("DETAIL_RETRIES", "2"))
    block_assets = os.environ.get("BLOCK_ASSETS", "1") != "0"

    # сколько дополнительно ждать хлебные крошки/даташит после появления Package/Description
    breadcrumb_wait = float(os.environ.get("BREADCRUMB_WAIT", "2.0"))

    # fallback кликом для даташита (0/1)
    datasheet_click_fallback = os.environ.get("DATASHEET_CLICK_FALLBACK", "1") != "0"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(
            accept_downloads=True,  # нужно для download.url fallback
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
        )
        context.set_default_timeout(8000)

        if block_assets:
            def _route(route):
                rt = route.request.resource_type
                # stylesheet НЕ трогаем
                if rt in ("image", "font", "media"):
                    route.abort()
                else:
                    route.continue_()
            context.route("**/*", _route)

        page = context.new_page()
        page.goto(URL_LIST, wait_until="domcontentloaded", timeout=120000)
        wait_table_ready(page, timeout_ms=60000)

        before = get_results_remaining(page)
        log(f"Results remaining BEFORE Basic tick: {before}")

        set_parts_type_basic(page)
        after_tick = wait_results_change(page, before, timeout_s=15)
        log(f"Results remaining AFTER  Basic tick (settled): {after_tick}")

        click_apply(page)
        wait_table_ready(page, timeout_ms=60000)
        after_apply = get_results_remaining(page)
        log(f"Results remaining AFTER  Apply: {after_apply}")

        # Phase 1: собрать Code + Link
        all_items_by_code: Dict[str, Dict] = {}
        page_idx = 1
        while True:
            batch, stats = scrape_list_basic_parts(page)
            for it in batch:
                all_items_by_code[it["Code"]] = it

            log(
                f"List Page {page_idx}: got={len(batch)} rows_total={stats.get('rows_total')} "
                f"rows_with_link={stats.get('rows_with_link')} total_unique={len(all_items_by_code)}"
            )

            if not goto_next_page_list(page):
                break
            page_idx += 1
            if page_idx > 500:
                break

        items = list(all_items_by_code.values())
        total = len(items)
        log(f"Phase1 total unique: {total}")

        # Phase 2: enrich
        page_detail = context.new_page()
        enriched: List[Dict] = []

        t_start = time.perf_counter()
        for i, it in enumerate(items, start=1):
            code = it.get("Code", "")
            it2 = enrich_one_detail(
                page_detail,
                it,
                min_interval_s=min_interval,
                retries=retries,
                breadcrumb_wait_s=breadcrumb_wait,
                datasheet_click_fallback=datasheet_click_fallback,
            )

            elapsed = time.perf_counter() - t_start
            avg = elapsed / i
            eta = (total - i) * avg

            ok = ("_detail_error" not in it2) and bool(it2.get("Package", "")) and bool(it2.get("Description", ""))
            status = "OK" if ok else "FAIL"
            inst = it2.get("InStock", "")

            log(f"{i}/{total} {code} In Stock: {inst} ... {status} [ETA: {fmt_dur(eta)}]")
            enriched.append(it2)

        browser.close()

    enriched_sorted = sorted(enriched, key=lambda r: r.get("Code", ""))
    date_tag = time.strftime("%Y-%m-%d")  # 2025-12-24
    out_xlsx = f"{OUT_XLSX_BASE}_{date_tag}.xlsx"

    save_excel(enriched_sorted, out_xlsx)
    log(f"SAVED {out_xlsx} rows={len(enriched_sorted)}")


if __name__ == "__main__":
    main()
